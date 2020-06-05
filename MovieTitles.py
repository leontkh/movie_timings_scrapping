from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.styles.colors import BLUE
from bs4 import BeautifulSoup
import requests
import re
import datetime

#prepping workbook
wb = Workbook()
ws = wb.active

ws["A1"] = "Last Updated:"
ws["B1"] = datetime.datetime.now()
ws["A2"] = "Movie Title"
ws["B2"] = "Movie Rating"
ws["C2"] = "Movie Length"
ws["D2"] = "Language"
ws["E2"] = "Location"
ws["F2"] = "Dates"
ws["G2"] = "Timings"
writingCellRow = 3

#Cathay Cineplexes
#Accessing the webpage
page = requests.get("http://www.cathaycineplexes.com.sg/movies/")
soup = BeautifulSoup(page.content, 'html.parser')



movies = soup.find_all(attrs = {"class":"boxgrid captionfull"})
for movie in movies:
    link = movie.find("a")["href"]
    link = "http://www.cathaycineplexes.com.sg" + link
    subpage = requests.get(link)
    subsoup = BeautifulSoup(subpage.content, 'html.parser')
    title = subsoup.find(id = "ContentPlaceHolder1_lblTitleM").get_text()
    lang = subsoup.find(id = "ContentPlaceHolder1_lblLanguage").get_text()
    rating = subsoup.find(id = "ContentPlaceHolder1_lblRating").get_text()
    time = subsoup.find(id = "ContentPlaceHolder1_lblRuntime").get_text()
    showpage = subsoup.find(id = "showtimes")
    for x in range(5,len(list(showpage.children))-3,2):
        loctimes = list(showpage.children)[x]
        loc = loctimes.find(class_ = "M_movietitle mobile").get_text()
        if loc != '':
            showtime = loctimes.find_all(class_ = "showtimeitem_time_pms")
            for timings in showtime:
                daytime = timings.find("a").attrs['title']
                pattern = r"(^.+) (.+) (.+$)"
                datetimesplit = re.match(pattern,daytime)
                ws["A"+str(count)] = '=HYPERLINK("{}", "{}")'.format(link, title)
                ws["B"+str(count)] = rating
                ws["C"+str(count)] = time
                ws["D"+str(count)] = lang
                ws["E"+str(count)] = loc
                ws["F"+str(count)] = datetimesplit.group(1)
                ws["G"+str(count)] = datetimesplit.group(2)+datetimesplit.group(3)
                writingCellRow += 1

#Retrieving date for the excel sheet
today=datetime.date.today()
#Saving excel sheet, with name that has a date in it
try:        
    wb.save(str(today)+"CathayMovieTimings.xlsx")
except:
    print("File is unable to save. Please check if file was left open.")
